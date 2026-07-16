import os, re, datetime, sqlite3, json, streamlit as st
from io import BytesIO
from openpyxl import Workbook
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter

st.set_page_config(page_title="Расчёт Заказов", page_icon="⚙️")
DAYS = ['Понедельник', 'Вторник', 'Среда', 'Четверг', 'Пятница', 'Суббота', 'Воскресенье']

# Инициализация сессии
if 'storage' not in st.session_state: 
    st.session_state.storage = []

def expand_serial_input(text):
    text = text.strip()
    if text.lower() == 'today': return True, DAYS[datetime.datetime.now().weekday()], 1
    parts = [p.strip() for p in re.split(r'[\s,]+', text) if p.strip()]
    if not parts: return False, "Строка пуста", 0
    count, res = 0, []
    for p in parts:
        if '-' in p:
            sub = p.split('-')
            s, e = sub[0].strip(), sub[1].strip()
            if len(e) < len(s): e = s[:len(s) - len(e)] + e
            count += int(e) - int(s) + 1
            res.append(f"{s}-{e}")
        elif p.isdigit(): 
            count += 1
            res.append(str(int(p)))
        else: 
            return False, f"Ошибка в: '{p}'", 0
    return True, ", ".join(res), count

def generate_excel_bytes(data):
    wb = Workbook()
    ws = wb.active
    ws.title = "Отчет за день"
    ws.append(["наименование", "номер чертежа", "номер операции", "стоимость за единицу", "номера изделий", "количество", "общая стоимость (операция)", "общая сумма за смену"])
    for c in range(1, 9): ws.cell(row=1, column=c).font = Font(bold=True)
    l_name, l_draw = "", ""
    for i in data:
        f_op = f"{i['op_num']} {i['desc']}"
        same = i['name'].lower() == l_name.lower() and i['drawing'] == l_draw
        ws.append(["" if same else i['name'], "" if same else i['drawing'], f_op, f"{i['price']:.2f} руб.", i['serials'], i['count'], f"{i['total']:.2f} руб.", ""])
        l_name, l_draw = i['name'], i['drawing']
    ws.cell(row=2, column=8).value = f"{sum(i['total'] for i in data):.2f} руб."
    for c in range(1, 9):
        ws.column_dimensions[get_column_letter(c)].width = max(max([len(str(ws.cell(row=r, column=c).value or '')) for r in range(1, ws.max_row + 1)]) + 4, 12)
    f = BytesIO()
    wb.save(f)
    return f.getvalue()

grand_total_now = sum(i['total'] for i in st.session_state.storage)
header_col, metric_col = st.columns(2)
with header_col: st.title("⚙️ Расчёт заказов")
with metric_col: st.metric(label="Сумма за смену", value=f"{grand_total_now:,.2f} руб.")

if not os.path.exists('production.db'):
    st.error("Файл 'production.db' не найден!")
else:
    with sqlite3.connect('production.db') as conn:
        db_names = [r[0] for r in conn.execute("SELECT DISTINCT name FROM items").fetchall()]

    import streamlit.components.v1 as components
    js_items = json.dumps(db_names)
    
    html_form = f"""
    <div style="font-family: sans-serif; width: 100%; box-sizing: border-box; padding: 5px;">
        <form id="main_work_form" onsubmit="sendData(event)">
            <div style="position: relative; margin-bottom: 15px;">
                <label style="font-weight: bold; font-size: 14px; color: #31333F;">Изделие:</label>
                <input type="text" id="inp_item" autocomplete="off" required style="width:100%; padding:10px; border:1px solid #ccc; border-radius:4px; font-size:16px; margin-top:5px; box-sizing:border-box;">
                <div id="box_sug" style="position: absolute; top: 100%; left: 0; width: 100%; background: white; border: 1px solid #ccc; border-top: none; border-radius: 0 0 4px 4px; box-shadow: 0 4px 10px rgba(0,0,0,0.15); display: none; z-index: 999999; max-height: 180px; overflow-y: auto;"></div>
            </div>
            <div style="margin-bottom: 15px;">
                <label style="font-weight: bold; font-size: 14px; color: #31333F;">Операции:</label>
                <input type="text" id="inp_ops" required style="width:100%; padding:10px; border:1px solid #ccc; border-radius:4px; font-size:16px; margin-top:5px; box-sizing:border-box;">
            </div>
            <div style="margin-bottom: 20px;">
                <label style="font-weight: bold; font-size: 14px; color: #31333F;">Номера изделий:</label>
                <input type="text" id="inp_serials" required style="width:100%; padding:10px; border:1px solid #ccc; border-radius:4px; font-size:16px; margin-top:5px; box-sizing:border-box;">
            </div>
            <button type="submit" style="width: 100%; background-color: rgb(255, 75, 75); color: white; border: none; padding: 12px; font-size: 16px; font-weight: bold; border-radius: 4px; cursor: pointer;">➕ Рассчитать и добавить</button>
        </form>
    </div>

    <script>
        const items = {js_items};
        const inpItem = document.getElementById('inp_item');
        const boxSug = document.getElementById('box_sug');
        const inpOps = document.getElementById('inp_ops');
        const inpSerials = document.getElementById('inp_serials');

        inpItem.addEventListener('input', (e) => {{
            const val = e.target.value.toLowerCase().trim();
            boxSug.innerHTML = '';
            if (!val) {{ boxSug.style.display = 'none'; return; }}
            const m = items.filter(i => i.toLowerCase().startsWith(val));
            if (m.length) {{
                m.forEach(i => {{
                    const d = document.createElement('div'); d.textContent = i;
                    d.style.padding = '12px 10px'; d.style.cursor = 'pointer'; d.style.borderBottom = '1px solid #eee';
                    d.onmouseenter = () => d.style.background = '#f5f5f5';
                    d.onmouseleave = () => d.style.background = 'white';
                    d.onclick = () => {{ inpItem.value = i; boxSug.style.display = 'none'; }};
                    boxSug.appendChild(d);
                }});
                boxSug.style.display = 'block';
            }} else {{ boxSug.style.display = 'none'; }}
        }});

        document.addEventListener('click', (e) => {{ if (e.target !== inpItem) boxSug.style.display = 'none'; }});

        function sendData(e) {{
            e.preventDefault();
            
            // Безопасно кодируем значения полей формы
            const pItem = encodeURIComponent(inpItem.value);
            const pOps = encodeURIComponent(inpOps.value);
            const pSerials = encodeURIComponent(inpSerials.value);
            
            // Напрямую перезагружаем родительское окно со стабильными query-параметрами URL. 
            // Это решает проблему отправки данных на абсолютно любых устройствах (включая iPhone).
            window.parent.location.search = `?item=${{pItem}}&ops=${{pOps}}&serials=${{pSerials}}`;

            setTimeout(() => {{
                inpItem.value = '';
                inpOps.value = '';
                inpSerials.value = '';
            }}, 100);
        }}
    </script>
    """
    
    # Отображаем форму
    components.html(html_form, height=330)
    # ОБРАБОТКА ДАННЫХ С ИСПОЛЬЗОВАНИЕМ СТАБИЛЬНОГО СЛОВАРЯ QUERY_PARAMS
    # Считываем переданные из формы параметры
    q_item = st.query_params.get("item", "").strip()
    q_ops = st.query_params.get("ops", "").strip()
    q_serials = st.query_params.get("serials", "").strip()

    if q_item and q_ops and q_serials:
        try:
            selected_name = q_item
            ops_raw = q_ops
            serials_raw = q_serials
            
            # Предварительно очищаем параметры URL, чтобы добавление не зацикливалось при обычном F5
            st.query_params.clear()
            
            ok, serials, count = expand_serial_input(serials_raw)
            if not ok: 
                st.error(serials)
            else:
                ops = [o.strip() for o in ops_raw.split(',') if o.strip()]
                found = []
                with sqlite3.connect('production.db') as conn:
                    cursor = conn.cursor()
                    for op in ops:
                        cursor.execute(
                            "SELECT drawing_number, work_description, price_per_unit FROM items WHERE LOWER(name)=LOWER(?) AND (work_description LIKE ? OR work_description LIKE ? OR work_description=?)", 
                            (selected_name, f'{op},%', f'{op} %', op)
                        )
                        res = cursor.fetchone()
                        if res: 
                            found.append({
                                'op_num': op, 
                                'desc': re.sub(r'^\d+\s*,\s*', '', str(res[1])).strip(), 
                                'price': float(res[2]), 
                                'drawing': str(res[0])
                            })

                if not found: 
                    st.error(f"Операции {ops_raw} для '{selected_name}' не найдены в базе данных.")
                else:
                    for o in found:
                        st.session_state.storage.append({
                            'name': selected_name, 
                            'drawing': o['drawing'], 
                            'op_num': o['op_num'], 
                            'desc': o['desc'], 
                            'price': o['price'], 
                            'serials': serials, 
                            'count': count, 
                            'total': o['price'] * count
                        })
                    st.success("Успешно добавлено!")
                    st.rerun()
        except Exception as e:
            st.sidebar.error(f"Ошибка сохранения: {e}")

    # Отрисовка результатов текущей смены
    if st.session_state.storage:
        st.write("---")
        with st.expander("🔍 Подробнее", expanded=True):
            for i in st.session_state.storage: 
                st.write(f"**{i['name']}** | Оп. {i['op_num']} ({i['desc']}) | {i['count']} шт. (№ {i['serials']}) — *{i['total']:.2f} руб.*")
        
        excel_file = generate_excel_bytes(st.session_state.storage)
        st.download_button(
            "💾 Скачать отчет Excel на iPhone", 
            excel_file, 
            f"{datetime.datetime.now().strftime('%d.%m.%Y')}.xlsx", 
            "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet", 
            use_container_width=True
        )
        
        if st.button("🗑️ Сбросить смену", use_container_width=True):
            st.session_state.storage = []
            st.rerun()

    # Админ-панель для работы с БД
    st.write("---")
    with st.expander("🔐 Редактор базы данных"):
        if st.text_input("Пароль администратора:", type="password", key="adm_p") == "1234":
            add_name = st.text_input("Наименование:").strip()
            add_draw = st.text_input("Чертеж:").strip()
            add_desc = st.text_input("Описание (начните с номера операции, например: '10, Токарная'):").strip()
            add_price = st.number_input("Цена:", min_value=0.0, step=0.5)
            
            if st.button("💾 Сохранить в базу данных", use_container_width=True):
                if not add_name or not add_draw or not add_desc or add_price <= 0: 
                    st.error("Заполните корректно все поля!")
                else:
                    with sqlite3.connect('production.db') as conn: 
                        conn.execute(
                            "INSERT INTO items (name, drawing_number, work_description, price_per_unit) VALUES (?, ?, ?, ?)", 
                            (add_name, add_draw, add_desc, add_price)
                        )
                        conn.commit()
                    st.success("Успешно добавлено в базу данных!")
                    st.rerun()
